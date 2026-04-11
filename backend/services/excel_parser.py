from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime, date, time
from typing import List, Dict, Any


# 歷史派單欄位對應（對應 1121已排班.xlsx）
HISTORY_FIELD_MAP = {
    "日期":        "order_date",
    "客戶":        "customer_type",
    "個案名":      "customer_name",
    "身分證字號":  "id_no",
    "連絡電話":    "contact_phone",
    "1.補助 2.自費": "payment_type",
    "性質":        "trip_type",
    "客上":        "pickup_time",
    "地區":        "region",
    "出發地":      "departure",
    "目的地":      "destination",
    "客下":        "dropoff_time",
    "里程":        "mileage",
    "車資":        "fare",
    "陪同金額":    "companion_fee",
    "自付金額":    "self_pay",
    "車號":        "plate_no",
    "司機":        "driver_name",
    "車型/配件":   "vehicle_accessories",
    "補助餘額":    "subsidy_balance",
    "補助餘額 ":   "subsidy_balance",
    "身分資格":    "qualification",
    "備註":        "note",
    "填單人":      "form_filler",
    "等級":        "grade",
}

# 上傳訂單欄位對應（未派車訂單）
ORDER_FIELD_MAP = {
    "日期":        "order_date",
    "客戶":        "customer_type",
    "個案名":      "customer_name",
    "身分證字號":  "id_no",
    "連絡電話":    "contact_phone",
    "1.補助 2.自費": "payment_type",
    "性質":        "trip_type",
    "客上":        "pickup_time",
    "地區":        "region",
    "出發地":      "departure",
    "目的地":      "destination",
    "里程":        "mileage",
    "車資":        "fare",
    "陪同金額":    "companion_fee",
    "自付金額":    "self_pay",
    "補助餘額":    "subsidy_balance",
    "補助餘額 ":   "subsidy_balance",
    "身分資格":    "qualification",
    "備註":        "note",
    "填單人":      "form_filler",
    "等級":        "grade",
}


def _normalize_header(h) -> str:
    return h.strip() if h else ""


def _parse_roc_date(val) -> date | None:
    """民國年轉西元：1141121 → 2025-11-21"""
    # 處理 float（如 1141121.0）先轉 int
    if isinstance(val, float):
        val = int(val)
    s = str(val).strip().replace("/", "").replace("-", "")
    if len(s) == 7 and s.isdigit():
        try:
            year = int(s[:3]) + 1911
            month = int(s[3:5])
            day = int(s[5:7])
            return date(year, month, day)
        except ValueError:
            return None
    # 嘗試直接解析西元日期
    for fmt in ("%Y%m%d", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _coerce(field: str, val, row_idx: int, errors: list):
    if val is None or str(val).strip() in ("", "nan", "None"):
        return None
    if field == "order_date":
        return _parse_roc_date(val)
    if field in ("pickup_time", "dropoff_time"):
        return _parse_time(val)
    if field == "payment_type":
        try:
            return int(val)
        except (TypeError, ValueError):
            return None
    if field in ("mileage", "fare", "companion_fee", "self_pay"):
        try:
            return float(str(val).replace(",", ""))
        except (TypeError, ValueError):
            return None
    return str(val).strip()


def _parse_excel(file_bytes: bytes, field_map: dict):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [_normalize_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

    rows, errors = [], []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v for v in row if v is not None):
            continue
        record: Dict[str, Any] = {}
        for col_idx, val in enumerate(row):
            if col_idx >= len(headers):
                break
            field = field_map.get(headers[col_idx])
            if field:
                record[field] = _coerce(field, val, row_idx, errors)
        # 組合 scheduled_time = order_date + pickup_time
        d = record.get("order_date")
        t = record.get("pickup_time")
        if d and t:
            record["scheduled_time"] = datetime.combine(d, t)
        elif d:
            record["scheduled_time"] = datetime.combine(d, time(0, 0))
        rows.append(record)
    return rows, errors


def parse_history_excel(file_bytes: bytes):
    return _parse_excel(file_bytes, HISTORY_FIELD_MAP)


def parse_orders_excel(file_bytes: bytes):
    return _parse_excel(file_bytes, ORDER_FIELD_MAP)


# 範本欄位（與 1121已排班.xlsx 相同）
_HISTORY_HEADERS = [
    "日期", "客戶", "個案名", "身分證字號", "連絡電話",
    "1.補助 2.自費", "性質", "客上", "地區", "出發地", "目的地", "客下",
    "里程", "車資", "陪同金額", "自付金額",
    "車號", "司機", "車型/配件", "補助餘額", "身分資格", "備註", "填單人", "等級"
]

_ORDER_HEADERS = [
    "日期", "客戶", "個案名", "身分證字號", "連絡電話",
    "1.補助 2.自費", "性質", "客上", "地區", "出發地", "目的地",
    "里程", "車資", "陪同金額", "自付金額", "補助餘額", "身分資格", "備註", "填單人", "等級"
]


def generate_history_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "歷史派單"
    ws.append(_HISTORY_HEADERS)
    ws.append([
        "1141121", "電話", "王小明", "A123456789", "0912-345-678",
        1, "送", "08:00", "北區", "台北市中正區XX路1號", "台大醫院", "10:00",
        5.2, 200, 0, 0,
        "ABC-1234", "張三", "一般", "1000-200=800", "低收入戶", "", "李四", ""
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_order_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "未派車訂單"
    ws.append(_ORDER_HEADERS)
    ws.append([
        "1141121", "電話", "陳小華", "B234567890", "0922-111-222",
        1, "送", "09:00", "北區", "台北市信義區YY路2號", "榮總醫院",
        3.5, 150, 0, 0, "", "低收入戶", "", "王五", ""
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
